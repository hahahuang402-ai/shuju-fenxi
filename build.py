import sys, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0,'/tmp')
import engine, core
from core import *
from wr import *
from engine import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string as C

wb = Workbook()
DV={'大中VAN':'大VAN'}; ZV={'大中VAN':'中VAN'}
TR={'燃料种类2':'传统燃料'}; NE={'燃料种类2':'新能源'}

# ============================================================ SHEET1
def sheet1(ws):
    cs = [
        (C('B'),'销量',{}),(C('C'),'同比',{}),(C('D'),'占有率',{}),(C('E'),'占同比',{}),
        (C('F'),'销量',DV),(C('G'),'同比',DV),(C('H'),'占有率',DV),(C('I'),'占同比',DV),
        (C('J'),'销量',{**DV,**TR}),(C('K'),'占有率',{**DV,**TR}),(C('L'),'占同比',{**DV,**TR}),
        (C('M'),'销量',{**DV,**NE}),(C('N'),'占有率',{**DV,**NE}),(C('O'),'占同比',{**DV,**NE}),
        (C('P'),'销量',ZV),(C('Q'),'同比',ZV),(C('R'),'占有率',ZV),(C('S'),'占同比',ZV),
        (C('T'),'销量',{**ZV,**TR}),(C('U'),'占有率',{**ZV,**TR}),(C('V'),'占同比',{**ZV,**TR}),
        (C('W'),'销量',{**ZV,**NE}),(C('X'),'占有率',{**ZV,**NE}),(C('Y'),'占同比',{**ZV,**NE}),
        (C('Z'),'销量',TR),(C('AA'),'同比',TR),(C('AB'),'占有率',TR),
        (C('AC'),'销量',NE),(C('AD'),'同比',NE),(C('AE'),'占有率',NE),
    ]
    ncol=C('AE'); ps=pct_set(cs); set_ztb(ztb_set(cs))
    setcell(ws,'A1','2026年1-4月销量',merge='A1:AE1')
    setcell(ws,'A2','企业品牌',merge='A2:A4'); setcell(ws,'B2','汇总',merge='B2:E3')
    setcell(ws,'F2','大VAN',merge='F2:O2'); setcell(ws,'P2','中VAN',merge='P2:Y2')
    setcell(ws,'Z2','传统燃料',merge='Z2:AB3'); setcell(ws,'AC2','新能源',merge='AC2:AE3')
    setcell(ws,'F3','大VAN汇总',merge='F3:I3'); setcell(ws,'J3','传统燃料',merge='J3:L3'); setcell(ws,'M3','新能源',merge='M3:O3')
    setcell(ws,'P3','中VAN汇总',merge='P3:S3'); setcell(ws,'T3','传统燃料',merge='T3:V3'); setcell(ws,'W3','新能源',merge='W3:Y3')
    h4=['销量','同比','占有率','占同比','销量','同比','占有率','占同比','销量','占有率','占同比','销量','占有率','占同比','销量','同比','占有率','占同比','销量','占有率','占同比','销量','占有率','占同比','销量','同比','占有率','销量','同比','占有率']
    for i,v in enumerate(h4): setcell(ws,f'{get_column_letter(2+i)}4',v,fill=HDR2)
    style_header_block(ws,1,4,1,ncol)
    r=5
    for b in brands_by_sales(CUR):
        write_row(ws,r,[(1,b)],build_values(cs,{},b),ps,ncol); r+=1
    write_row(ws,r,[(1,'总计')],build_values(cs,{},None),ps,ncol,fill=TOT,bold=True)
    set_widths(ws,11,8.5,ncol); ws.freeze_panes='B5'

# ============================================================ SHEET2  省份->品牌
def sheet2(ws):
    cs = [
        (C('C'),'销量',{}),(C('D'),'同比',{}),(C('E'),'占有率',{}),(C('F'),'占同比',{}),
        (C('G'),'销量',DV),(C('H'),'同比',DV),(C('I'),'占有率',DV),(C('J'),'占同比',DV),
        (C('K'),'销量',{**DV,**TR}),(C('L'),'同比',{**DV,**TR}),(C('M'),'占有率',{**DV,**TR}),(C('N'),'占同比',{**DV,**TR}),
        (C('O'),'销量',{**DV,**NE}),(C('P'),'同比',{**DV,**NE}),(C('Q'),'占有率',{**DV,**NE}),(C('R'),'占同比',{**DV,**NE}),
        (C('S'),'销量',ZV),(C('T'),'同比',ZV),(C('U'),'占有率',ZV),(C('V'),'占同比',ZV),
        (C('W'),'销量',{**ZV,**TR}),(C('X'),'同比',{**ZV,**TR}),(C('Y'),'占有率',{**ZV,**TR}),(C('Z'),'占同比',{**ZV,**TR}),
        (C('AA'),'销量',{**ZV,**NE}),(C('AB'),'同比',{**ZV,**NE}),(C('AC'),'占有率',{**ZV,**NE}),(C('AD'),'占同比',{**ZV,**NE}),
    ]
    ncol=C('AD'); ps=pct_set(cs); set_ztb(ztb_set(cs))
    setcell(ws,'A1','2026年1-4月销量',merge='A1:AD1')
    setcell(ws,'A2','省份',merge='A2:A4'); setcell(ws,'B2','企业品牌',merge='B2:B4'); setcell(ws,'C2','汇总',merge='C2:F3')
    setcell(ws,'G2','大VAN',merge='G2:R2'); setcell(ws,'S2','中VAN',merge='S2:AD2')
    setcell(ws,'G3','大VAN汇总',merge='G3:J3'); setcell(ws,'K3','传统燃料',merge='K3:N3'); setcell(ws,'O3','新能源',merge='O3:R3')
    setcell(ws,'S3','中VAN汇总',merge='S3:V3'); setcell(ws,'W3','传统燃料',merge='W3:Z3'); setcell(ws,'AA3','新能源',merge='AA3:AD3')
    h4=['销量','同比','占有率','占同比']*7
    for i,v in enumerate(h4): setcell(ws,f'{get_column_letter(3+i)}4',v,fill=HDR2)
    style_header_block(ws,1,4,1,ncol)
    r=5
    for prov in ['甘肃','内蒙古','宁夏','青海','陕西']:
        pf={'省份':prov}; bs=brands_by_sales(CUR,pf); rstart=r
        for b in bs:
            write_row(ws,r,[(2,b)],build_values(cs,pf,b,denom_rgf=pf),ps,ncol); r+=1
        write_row(ws,r,[(1,f'{prov} 汇总')],build_values(cs,pf,None,denom_rgf=pf),ps,ncol,fill=SUBT,bold=True)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        if bs:
            ws.merge_cells(start_row=rstart,start_column=1,end_row=rstart+len(bs)-1,end_column=1)
            ws.cell(row=rstart,column=1,value=prov).alignment=CEN
        r+=1
    write_row(ws,r,[(1,'总计')],build_values(cs,{},None),ps,ncol,fill=TOT,bold=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    set_widths(ws,10,8.2,ncol,first_cols=2); ws.freeze_panes='C5'

# ============================================================ SHEET3  省份 x 福田
def sheet3(ws):
    FT={'brand':'福田'}
    cs=[
        (C('B'),'销量',{}),(C('C'),'同比',{}),
        (C('D'),'销量',{},FT),(C('E'),'占有率',{},FT),(C('F'),'占同比',{},FT),
        (C('G'),'销量',DV),(C('H'),'同比',DV),
        (C('I'),'销量',DV,FT),(C('J'),'占有率',DV,FT),(C('K'),'占同比',DV,FT),
        (C('L'),'销量',{**DV,**TR}),(C('M'),'销量',{**DV,**TR},FT),(C('N'),'占有率',{**DV,**TR},FT),(C('O'),'占同比',{**DV,**TR},FT),
        (C('P'),'销量',{**DV,**NE}),(C('Q'),'销量',{**DV,**NE},FT),(C('R'),'占有率',{**DV,**NE},FT),(C('S'),'占同比',{**DV,**NE},FT),
        (C('T'),'销量',ZV),(C('U'),'同比',ZV),
        (C('V'),'销量',ZV,FT),(C('W'),'占有率',ZV,FT),(C('X'),'占同比',ZV,FT),
        (C('Y'),'销量',{**ZV,**TR}),(C('Z'),'销量',{**ZV,**TR},FT),(C('AA'),'占有率',{**ZV,**TR},FT),(C('AB'),'占同比',{**ZV,**TR},FT),
        (C('AC'),'销量',{**ZV,**NE}),(C('AD'),'销量',{**ZV,**NE},FT),(C('AE'),'占有率',{**ZV,**NE},FT),(C('AF'),'占同比',{**ZV,**NE},FT),
    ]
    ncol=C('AF'); ps=pct_set(cs); set_ztb(ztb_set(cs))
    setcell(ws,'A1','2026年1-4月销量',merge='A1:AF1')
    setcell(ws,'A2','省份',merge='A2:A5'); setcell(ws,'B2','汇总',merge='B2:C4'); setcell(ws,'D2','福田',merge='D2:F4')
    setcell(ws,'G2','大VAN',merge='G2:S2'); setcell(ws,'T2','中VAN',merge='T2:AF2')
    setcell(ws,'G3','大VAN汇总',merge='G3:H4'); setcell(ws,'I3','福田大VAN',merge='I3:K4')
    setcell(ws,'L3','传统燃料',merge='L3:O3'); setcell(ws,'P3','新能源',merge='P3:S3')
    setcell(ws,'T3','中VAN汇总',merge='T3:U4'); setcell(ws,'V3','福田中VAN',merge='V3:X4')
    setcell(ws,'Y3','传统燃料',merge='Y3:AB3'); setcell(ws,'AC3','新能源',merge='AC3:AF3')
    for col,txt in [('L4','汇总'),('M4','福田'),('P4','汇总'),('Q4','福田'),('Y4','汇总'),('Z4','福田'),('AC4','汇总'),('AD4','福田')]:
        setcell(ws,col,txt)
    h5=['销量','同比','销量','占有率','占同比','销量','同比','销量','占有率','占同比','销量','销量','占有率','占同比','销量','销量','占有率','占同比','销量','同比','销量','占有率','占同比','销量','销量','占有率','占同比','销量','销量','占有率','占同比']
    for i,v in enumerate(h5): setcell(ws,f'{get_column_letter(2+i)}5',v,fill=HDR2)
    style_header_block(ws,1,5,1,ncol)
    r=6
    for prov in ['陕西','内蒙古','甘肃','宁夏','青海']:
        pf={'省份':prov}
        write_row(ws,r,[(1,prov)],build_values(cs,pf,None,denom_rgf=pf),ps,ncol); r+=1
    write_row(ws,r,[(1,'总计')],build_values(cs,{},None),ps,ncol,fill=TOT,bold=True)
    set_widths(ws,10,8.2,ncol); ws.freeze_panes='B6'

# ============================================================ SHEET4  省份->城市->燃料  (含比重)
def sheet4(ws):
    setcell(ws,'A1','2026年1-4月销量',merge='A1:AA1')
    setcell(ws,'A2','省份',merge='A2:A4'); setcell(ws,'B2','城市',merge='B2:B4'); setcell(ws,'C2','燃料种类',merge='C2:C4')
    setcell(ws,'D2','整体市场情况',merge='D2:K2'); setcell(ws,'L2','大VAN',merge='L2:S2'); setcell(ws,'T2','中VAN',merge='T2:AA2')
    setcell(ws,'D3','汇总',merge='D3:F3'); setcell(ws,'G3','福田',merge='G3:K3')
    setcell(ws,'L3','大VAN汇总',merge='L3:N3'); setcell(ws,'O3','福田',merge='O3:S3')
    setcell(ws,'T3','中VAN汇总',merge='T3:V3'); setcell(ws,'W3','福田',merge='W3:AA3')
    h4=['销量','同比','比重','销量','同比','比重','占有率','占同比']*3
    ncol=C('AA')
    for i,v in enumerate(h4): setcell(ws,f'{get_column_letter(4+i)}4',v,fill=HDR2)
    style_header_block(ws,1,4,1,ncol)
    def specfor(parent_rgf):
        FT={'brand':'福田'}
        spec=[]
        for cbase,scope in [(C('D'),{}),(C('L'),DV),(C('T'),ZV)]:
            layout=[(0,'销量',{}),(1,'同比',{}),(2,'比重',{'denom':parent_rgf}),
                    (3,'销量',FT),(4,'同比',FT),(5,'比重',{**FT,'denom':parent_rgf}),
                    (6,'占有率',FT),(7,'占同比',FT)]
            for off,want,opt in layout:
                spec.append((cbase+off,want,scope,opt))
        return spec
    ps={C('F'),C('I'),C('J'),C('K'),C('N'),C('Q'),C('R'),C('S'),C('V'),C('Y'),C('Z'),C('AA'),C('E'),C('H'),C('M'),C('P'),C('U'),C('X')}
    # rebuild ps from a sample spec
    ps=pct_set(specfor({})); set_ztb(ztb_set(specfor({})))
    prov_cities={
        '甘肃':['兰州市','酒泉市','张掖市','定西市','平凉市','庆阳市','天水市','武威市','白银市'],
        '内蒙古':['呼和浩特市','包头市','鄂尔多斯市','巴彦淖尔市','乌海市','阿拉善盟'],
        '宁夏':['银川市','固原市','石嘴山市','吴忠市','中卫市'],
        '青海':['西宁市','海东地区','玉树','果洛','海北','海南','海西','黄南'],
        '陕西':['西安市','渭南市','榆林市','咸阳市','汉中市','宝鸡市','安康市','延安市','商洛市','铜川市'],
    }
    r=5
    for prov,cities in prov_cities.items():
        provf={'省份':prov}; prov_rs=r
        for city in cities:
            cityf={'省份':prov,'城市':city}; city_rs=r
            for fname,ff in [('传统燃料',TR),('新能源',NE)]:
                rgf={**cityf,**ff}
                if ssum(CUR,rgf)<=0 and ssum(PRE,rgf)<=0: continue
                write_row(ws,r,[(3,fname)],build_values(specfor(cityf),rgf,None),ps,ncol); r+=1
            write_row(ws,r,[(2,f'{city} 汇总')],build_values(specfor(provf),cityf,None),ps,ncol,fill=SUBT,bold=True)
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
            if r>city_rs:
                ws.merge_cells(start_row=city_rs,start_column=2,end_row=r-1,end_column=2)
                ws.cell(row=city_rs,column=2,value=city).alignment=CEN
            r+=1
        write_row(ws,r,[(1,f'{prov} 汇总')],build_values(specfor({}),provf,None),ps,ncol,fill=TOT,bold=True)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        ws.merge_cells(start_row=prov_rs,start_column=1,end_row=r-1,end_column=1)
        ws.cell(row=prov_rs,column=1,value=prov).alignment=CEN
        r+=1
    write_row(ws,r,[(1,'总计')],build_values(specfor({}),{},None),ps,ncol,fill=TOT,bold=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    set_widths(ws,9,8.0,ncol,first_cols=3); ws.freeze_panes='D5'
# === sheets 5-9 appended below ===

# ============================================================ SHEET5  大VAN 燃料->米段->品牌, 分场景
def sheet5(ws):
    engine.BASE = DV
    scenes=[('改装定制',C('J')),('客货两用',C('N')),('商旅客运',C('R')),('专业物流',C('V'))]
    def specfor(denom_rgf):
        cs=[
            (C('D'),'销量',{},{'curF':Y25}),(C('E'),'占有率',{},{'curF':Y25,'denom':denom_rgf}),
            (C('F'),'销量',{}),(C('G'),'同比',{}),(C('H'),'占有率',{},{'denom':denom_rgf}),(C('I'),'占同比',{},{'denom':denom_rgf}),
        ]
        for sc,cb in scenes:
            ff={'功能分类2':sc}
            cs += [(cb,'销量',ff),(cb+1,'同比',ff),(cb+2,'占有率',ff,{'denom':denom_rgf}),(cb+3,'占同比',ff,{'denom':denom_rgf})]
        return cs
    ncol=C('Y'); ps=pct_set(specfor({})); set_ztb(ztb_set(specfor({})))
    setcell(ws,'A1','2026年1-4月大VAN销量',merge='A1:Y1')
    setcell(ws,'A2','分燃料',merge='A2:A4'); setcell(ws,'B2','米段',merge='B2:B4'); setcell(ws,'C2','企业品牌',merge='C2:C4')
    setcell(ws,'D2','2025年全年',merge='D2:E3'); setcell(ws,'F2','2026年1-4月汇总',merge='F2:I3'); setcell(ws,'J2','分场景',merge='J2:Y2')
    setcell(ws,'J3','改装定制',merge='J3:M3'); setcell(ws,'N3','客货两用',merge='N3:Q3'); setcell(ws,'R3','商旅客运',merge='R3:U3'); setcell(ws,'V3','专业物流',merge='V3:Y3')
    h4=['销量','占有率','销量','同比','占有率','占同比']+['销量','同比','占有率','占同比']*4
    for i,v in enumerate(h4): setcell(ws,f'{get_column_letter(4+i)}4',v,fill=HDR2)
    style_header_block(ws,1,4,1,ncol)
    r=5
    for fuel,ff in [('传统燃料',TR),('新能源',NE)]:
        fuel_rs=r
        msegs=[m for m in ['5.0米','5.5米','6.0米'] if ssum(CUR,{**ff,MSEG:m})>0]
        for mseg in msegs:
            grp={**ff,MSEG:mseg}; m_rs=r
            bs=brands_by_sales(CUR,grp,base=DV)
            for b in bs:
                write_row(ws,r,[(3,b)],build_values(specfor(grp),grp,b),ps,ncol); r+=1
            write_row(ws,r,[(2,f'{mseg} 汇总')],build_values(specfor(grp),grp,None),ps,ncol,fill=SUBT,bold=True)
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
            if bs:
                ws.merge_cells(start_row=m_rs,start_column=2,end_row=m_rs+len(bs)-1,end_column=2)
                ws.cell(row=m_rs,column=2,value=mseg).alignment=CEN
            r+=1
        write_row(ws,r,[(1,f'{fuel} 汇总')],build_values(specfor(ff),ff,None),ps,ncol,fill=TOT,bold=True)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        ws.merge_cells(start_row=fuel_rs,start_column=1,end_row=r-1,end_column=1)
        ws.cell(row=fuel_rs,column=1,value=fuel).alignment=CEN
        r+=1
    write_row(ws,r,[(1,'总计')],build_values(specfor({}),{},None),ps,ncol,fill=TOT,bold=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    set_widths(ws,9,8.0,ncol,first_cols=3); ws.freeze_panes='D5'
    engine.BASE={}

# ============================================================ SHEET6  大VAN 米段x功能
def sheet6(ws):
    engine.BASE = DV
    setcell(ws,'A1','2026年1-4月大VAN市场销量',merge='A1:AH1')
    setcell(ws,'A2','米段',merge='A2:A4'); setcell(ws,'B2','功能分类',merge='B2:B4'); setcell(ws,'C2','汇总',merge='C2:D3')
    setcell(ws,'E2','传统燃料',merge='E2:V2'); setcell(ws,'W2','新能源',merge='W2:AH2')
    setcell(ws,'E3','传统燃料汇总',merge='E3:F3'); setcell(ws,'G3','全顺',merge='G3:I3'); setcell(ws,'J3','依维柯',merge='J3:L3')
    setcell(ws,'M3','大通',merge='M3:O3'); setcell(ws,'P3','福田',merge='P3:S3'); setcell(ws,'T3','重汽',merge='T3:V3')
    setcell(ws,'W3','新能源汇总',merge='W3:X3'); setcell(ws,'Y3','吉利',merge='Y3:AA3'); setcell(ws,'AB3','依维柯',merge='AB3:AD3'); setcell(ws,'AE3','福田',merge='AE3:AH3')
    h4=['销量','比重','销量','比重','销量','比重','占有率','销量','比重','占有率','销量','比重','占有率','销量','比重','占有率','占同比','销量','比重','占有率',
        '销量','比重','销量','比重','占有率','销量','比重','占有率','销量','比重','占有率','占同比']
    ncol=C('AH')
    for i,v in enumerate(h4): setcell(ws,f'{get_column_letter(3+i)}4',v,fill=HDR2)
    style_header_block(ws,1,4,1,ncol)
    def specfor(parent_rgf):
        return [
            (C('C'),'销量',{}),(C('D'),'比重',{},{'denom':parent_rgf}),
            (C('E'),'销量',TR),(C('F'),'比重',TR,{'denom':parent_rgf}),
            (C('G'),'销量',TR,{'brand':'全顺'}),(C('H'),'比重',TR,{'brand':'全顺','denom':parent_rgf}),(C('I'),'占有率',TR,{'brand':'全顺'}),
            (C('J'),'销量',TR,{'brand':'依维柯'}),(C('K'),'比重',TR,{'brand':'依维柯','denom':parent_rgf}),(C('L'),'占有率',TR,{'brand':'依维柯'}),
            (C('M'),'销量',TR,{'brand':'大通'}),(C('N'),'比重',TR,{'brand':'大通','denom':parent_rgf}),(C('O'),'占有率',TR,{'brand':'大通'}),
            (C('P'),'销量',TR,{'brand':'福田'}),(C('Q'),'比重',TR,{'brand':'福田','denom':parent_rgf}),(C('R'),'占有率',TR,{'brand':'福田'}),(C('S'),'占同比',TR,{'brand':'福田'}),
            (C('T'),'销量',TR,{'brand':'重汽'}),(C('U'),'比重',TR,{'brand':'重汽','denom':parent_rgf}),(C('V'),'占有率',TR,{'brand':'重汽'}),
            (C('W'),'销量',NE),(C('X'),'比重',NE,{'denom':parent_rgf}),
            (C('Y'),'销量',NE,{'brand':'吉利'}),(C('Z'),'比重',NE,{'brand':'吉利','denom':parent_rgf}),(C('AA'),'占有率',NE,{'brand':'吉利'}),
            (C('AB'),'销量',NE,{'brand':'依维柯'}),(C('AC'),'比重',NE,{'brand':'依维柯','denom':parent_rgf}),(C('AD'),'占有率',NE,{'brand':'依维柯'}),
            (C('AE'),'销量',NE,{'brand':'福田'}),(C('AF'),'比重',NE,{'brand':'福田','denom':parent_rgf}),(C('AG'),'占有率',NE,{'brand':'福田'}),(C('AH'),'占同比',NE,{'brand':'福田'}),
        ]
    ps=pct_set(specfor({})); set_ztb(ztb_set(specfor({})))
    funcs=['改装定制','客货两用','商旅客运','专业物流']
    r=5
    for mseg in ['5.0米','5.5米','6.0米']:
        mf={MSEG:mseg}; m_rs=r
        for fn in funcs:
            grp={MSEG:mseg,'功能分类2':fn}
            write_row(ws,r,[(2,fn)],build_values(specfor(mf),grp,None),ps,ncol); r+=1
        write_row(ws,r,[(1,f'{mseg} 汇总')],build_values(specfor({}),mf,None),ps,ncol,fill=SUBT,bold=True)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        ws.merge_cells(start_row=m_rs,start_column=1,end_row=m_rs+len(funcs)-1,end_column=1)
        ws.cell(row=m_rs,column=1,value=mseg).alignment=CEN
        r+=1
    all_rs=r
    for fn in funcs:
        grp={'功能分类2':fn}
        write_row(ws,r,[(2,fn)],build_values(specfor({}),grp,None),ps,ncol); r+=1
    write_row(ws,r,[(1,'总计')],build_values(specfor({}),{},None),ps,ncol,fill=TOT,bold=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    ws.merge_cells(start_row=all_rs,start_column=1,end_row=r-1,end_column=1)
    ws.cell(row=all_rs,column=1,value='所有米段汇总').alignment=CEN
    set_widths(ws,9,7.5,ncol,first_cols=2); ws.freeze_panes='C5'
    engine.BASE={}

# ============================================================ SHEET7  中VAN 明盲窗->品牌
def sheet7(ws):
    engine.BASE = ZV
    V6={VOL:'6m³'}; V7={VOL:'7m³'}; V8={VOL:'8m³'}
    def specfor(denom_rgf):
        return [
            (C('C'),'销量',{}),(C('D'),'同比',{}),(C('E'),'占有率',{},{'denom':denom_rgf}),(C('F'),'占同比',{},{'denom':denom_rgf}),
            (C('G'),'销量',TR),(C('H'),'同比',TR),(C('I'),'占有率',TR,{'denom':denom_rgf}),(C('J'),'占同比',TR,{'denom':denom_rgf}),
            (C('K'),'销量',NE),(C('L'),'同比',NE),(C('M'),'占有率',NE,{'denom':denom_rgf}),(C('N'),'占同比',NE,{'denom':denom_rgf}),
            (C('O'),'销量',{**NE,**V6}),(C('P'),'同比',{**NE,**V6}),(C('Q'),'占有率',{**NE,**V6},{'denom':denom_rgf}),(C('R'),'占同比',{**NE,**V6},{'denom':denom_rgf}),
            (C('S'),'销量',{**NE,**V7}),(C('T'),'同比',{**NE,**V7}),(C('U'),'占有率',{**NE,**V7},{'denom':denom_rgf}),(C('V'),'占同比',{**NE,**V7},{'denom':denom_rgf}),
            (C('W'),'销量',{**NE,**V8}),(C('X'),'同比',{**NE,**V8}),(C('Y'),'占有率',{**NE,**V8},{'denom':denom_rgf}),(C('Z'),'占同比',{**NE,**V8},{'denom':denom_rgf}),
        ]
    ncol=C('Z'); ps=pct_set(specfor({})); set_ztb(ztb_set(specfor({})))
    setcell(ws,'A1','2026年1-4月中VAN市场数据',merge='A1:Z1')
    setcell(ws,'A2','明盲窗',merge='A2:A4'); setcell(ws,'B2','企业品牌',merge='B2:B4'); setcell(ws,'C2','汇总',merge='C2:F3')
    setcell(ws,'G2','传统能源汇总',merge='G2:J3'); setcell(ws,'K2','新能源汇总',merge='K2:N3'); setcell(ws,'O2','新能源',merge='O2:Z2')
    setcell(ws,'O3','6m³',merge='O3:R3'); setcell(ws,'S3','7m³',merge='S3:V3'); setcell(ws,'W3','8m³',merge='W3:Z3')
    h4=['销量','同比','占有率','占同比']*6
    for i,v in enumerate(h4): setcell(ws,f'{get_column_letter(3+i)}4',v,fill=HDR2)
    style_header_block(ws,1,4,1,ncol)
    r=5
    ov_rs=r; bs=brands_by_sales(CUR,base=ZV)
    for b in bs:
        write_row(ws,r,[(2,b)],build_values(specfor({}),{},b),ps,ncol); r+=1
    write_row(ws,r,[(1,'总计')],build_values(specfor({}),{},None),ps,ncol,fill=TOT,bold=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    ws.merge_cells(start_row=ov_rs,start_column=1,end_row=ov_rs+len(bs)-1,end_column=1)
    ws.cell(row=ov_rs,column=1,value='整体').alignment=CEN
    r+=1
    for win in ['盲窗','明窗']:
        wf={'窗':win}; w_rs=r; bs=brands_by_sales(CUR,wf,base=ZV)
        for b in bs:
            write_row(ws,r,[(2,b)],build_values(specfor(wf),wf,b),ps,ncol); r+=1
        write_row(ws,r,[(1,f'{win} 汇总')],build_values(specfor(wf),wf,None),ps,ncol,fill=SUBT,bold=True)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        ws.merge_cells(start_row=w_rs,start_column=1,end_row=w_rs+len(bs)-1,end_column=1)
        ws.cell(row=w_rs,column=1,value=win).alignment=CEN
        r+=1
    set_widths(ws,9,8.0,ncol,first_cols=2); ws.freeze_panes='C5'
    engine.BASE={}

# ============================================================ SHEET8  风景G系列(非纯电中VAN)
def sheet8(ws):
    FUEL='燃料种类'
    VANS={'欧系中VAN':(C('C'),C('D')),'日系大VAN':(C('H'),C('I')),'日系中VAN':(C('M'),C('N'))}
    fuel_map={'CNG':['CNG','天然气'],'柴油':['柴油'],'纯电动':['纯电动'],'汽油':['汽油']}
    vanset=['欧系中VAN','日系大VAN','日系中VAN']
    def specfor(van, denom_rgf, fuelvals):
        c25, c26 = VANS[van]
        vf={'VAN种类':van, FUEL:fuelvals}
        return [
            (c25,'销量',vf,{'curF':Y25}),
            (c26,'销量',vf),(c26+1,'同比',vf),(c26+2,'占有率',vf,{'denom':denom_rgf}),(c26+3,'占同比',vf,{'denom':denom_rgf}),
        ]
    def full_spec(denom_rgf, fuelvals):
        spec=[]
        for van in vanset: spec+=specfor(van,denom_rgf,fuelvals)
        return spec
    ncol=C('Q'); ps=pct_set(full_spec({},['CNG'])); set_ztb(ztb_set(full_spec({},['CNG'])))
    setcell(ws,'A1','2026年1-4月风景G系列市场情况',merge='A1:Q1')
    setcell(ws,'A2','燃料种类',merge='A2:A4'); setcell(ws,'B2','企业品牌',merge='B2:B4')
    setcell(ws,'C2','欧系中VAN',merge='C2:G2'); setcell(ws,'H2','日系大VAN',merge='H2:L2'); setcell(ws,'M2','日系中VAN',merge='M2:Q2')
    setcell(ws,'C3','2025全年'); setcell(ws,'D3','2026年1-4月',merge='D3:G3')
    setcell(ws,'H3','2025全年'); setcell(ws,'I3','2026年1-4月',merge='I3:L3')
    setcell(ws,'M3','2025全年'); setcell(ws,'N3','2026年1-4月',merge='N3:Q3')
    h4=['销量','销量','同比','占有率','占同比','销量','销量','同比','占有率','占同比','销量','销量','同比','占有率','占同比']
    for i,v in enumerate(h4): setcell(ws,f'{get_column_letter(3+i)}4',v,fill=HDR2)
    style_header_block(ws,1,4,1,ncol)
    base3=CUR[CUR['VAN种类'].isin(vanset)]
    r=5
    for fname,fvals in fuel_map.items():
        ff={FUEL:fvals}; f_rs=r
        present=sub(base3,ff)
        bs=[b for b,v in present.groupby('企业品牌')['数量'].sum().sort_values(ascending=False).items() if v>0]
        if not bs and ssum(Y25,{**ff,'VAN种类':vanset})<=0:
            continue
        for b in bs:
            write_row(ws,r,[(2,b)],build_values(full_spec(ff,fvals), ff, b),ps,ncol); r+=1
        write_row(ws,r,[(1,f'{fname} 汇总')],build_values(full_spec(ff,fvals), ff, None),ps,ncol,fill=SUBT,bold=True)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        if bs:
            ws.merge_cells(start_row=f_rs,start_column=1,end_row=f_rs+len(bs)-1,end_column=1)
            ws.cell(row=f_rs,column=1,value=fname).alignment=CEN
        r+=1
    allvals=sum(fuel_map.values(),[]); allf={FUEL:allvals}
    write_row(ws,r,[(1,'总计')],build_values(full_spec(allf,allvals), allf, None),ps,ncol,fill=TOT,bold=True)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    set_widths(ws,9,8.0,ncol,first_cols=2); ws.freeze_panes='C5'

# ============================================================ SHEET9  10类专用
def sheet9(ws):
    K='10类专用'
    BR={'全顺':C('E'),'依维柯':C('J'),'大通':C('O'),'福田':C('T')}
    cats=['休闲旅居','工程作业','医疗卫生','运钞押运','警用特种','冷链运输','流动服务','民生服务','检测监测','邮政运输']
    cats=[c for c in cats if ssum(CUR,{K:c})>0 or ssum(PRE,{K:c})>0]
    allcats={K:cats}
    def specfor(parent_rgf):
        spec=[(C('B'),'销量',{}),(C('C'),'同比',{}),(C('D'),'比重',{},{'denom':parent_rgf})]
        for b,cb in BR.items():
            spec+=[(cb,'销量',{},{'brand':b}),(cb+1,'同比',{},{'brand':b}),
                   (cb+2,'比重',{},{'brand':b,'denom':parent_rgf}),
                   (cb+3,'占有率',{},{'brand':b}),(cb+4,'占同比',{},{'brand':b})]
        return spec
    ncol=C('X'); ps=pct_set(specfor({})); set_ztb(ztb_set(specfor({})))
    setcell(ws,'A1','2026年1-4月专用类市场情况',merge='A1:X1')
    setcell(ws,'A2','10类专用',merge='A2:A3'); setcell(ws,'B2','汇总',merge='B2:D2')
    setcell(ws,'E2','全顺',merge='E2:I2'); setcell(ws,'J2','依维柯',merge='J2:N2'); setcell(ws,'O2','大通',merge='O2:S2'); setcell(ws,'T2','福田',merge='T2:X2')
    h3=['销量','同比','比重']+['销量','同比','比重','占有率','占同比']*4
    for i,v in enumerate(h3): setcell(ws,f'{get_column_letter(2+i)}3',v,fill=HDR2)
    style_header_block(ws,1,3,1,ncol)
    r=4
    for cat in cats:
        write_row(ws,r,[(1,cat)],build_values(specfor(allcats),{K:cat},None),ps,ncol); r+=1
    write_row(ws,r,[(1,'总计')],build_values(specfor(allcats),allcats,None),ps,ncol,fill=TOT,bold=True)
    set_widths(ws,10,8.0,ncol); ws.freeze_panes='B4'

s1=wb.active; s1.title='sheet1'; sheet1(s1)
sheet2(wb.create_sheet('sheet2'))
sheet3(wb.create_sheet('sheet3'))
sheet4(wb.create_sheet('sheet4'))
sheet5(wb.create_sheet('sheet5'))
sheet6(wb.create_sheet('sheet6'))
sheet7(wb.create_sheet('sheet7'))
sheet8(wb.create_sheet('sheet8'))
sheet9(wb.create_sheet('sheet9'))
wb.save('26年1-4月分析结果.xlsx')
print('all 9 sheets saved')
