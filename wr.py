import warnings
warnings.filterwarnings('ignore')
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
HDR  = PatternFill('solid', fgColor='D9E1F2')   # 表头蓝
HDR2 = PatternFill('solid', fgColor='E2EFDA')    # 指标行绿
TOT  = PatternFill('solid', fgColor='FFF2CC')    # 小计/总计黄
SUBT = PatternFill('solid', fgColor='FCE4D6')    # 分类汇总橙
GREEN = PatternFill('solid', fgColor='C6EFCE')   # 占同比增长大-绿
RED   = PatternFill('solid', fgColor='FFC7CE')   # 占同比下降大-红
GREEN_FONT = Font(color='006100')
RED_FONT   = Font(color='9C0006')
ZTB_THRESH = float('inf')   # 禁用占同比条件着色(用户自行着色)
CUR_ZTB = set()    # 当前sheet的占同比列集合(由各sheet设置)
def set_ztb(cols):
    global CUR_ZTB; CUR_ZTB = cols
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
CEN = Alignment(horizontal='center', vertical='center')
BOLD = Font(bold=True)

def setcell(ws, coord, val, merge=None, fill=HDR, bold=True):
    c = ws[coord]; c.value = val
    if merge: ws.merge_cells(merge)
    c.alignment = CENTER; c.border = BORDER; c.font = Font(bold=bold)
    if fill: c.fill = fill
    return c

def style_header_block(ws, r1, r2, c1, c2, fill=HDR):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CENTER; cell.border = BORDER; cell.font = BOLD
            if fill: cell.fill = fill

def write_row(ws, r, label_cells, values, pct_idx, ncol, fill=None, bold=False, ztb_idx=None):
    """label_cells: list of (col, text). values: dict col->val. pct_idx: set of cols showing %.
    ztb_idx: set of 占同比 cols to conditionally color (green up / red down)."""
    ztb_idx = ztb_idx if ztb_idx is not None else CUR_ZTB
    for col, txt in label_cells:
        cell = ws.cell(row=r, column=col, value=txt)
        cell.alignment = CEN; cell.border = BORDER
        if bold: cell.font = BOLD
        if fill: cell.fill = fill
    for col in range(1, ncol+1):
        if any(col==lc[0] for lc in label_cells): continue
        cell = ws.cell(row=r, column=col)
        val = values.get(col)
        if val is None:
            cell.value = None
        elif val == '净增长':
            cell.value = '净增长'
        elif col in pct_idx:
            cell.value = val/100.0
            cell.number_format = '0.0%'
        else:
            cell.value = val
        cell.alignment = CEN; cell.border = BORDER
        if bold: cell.font = BOLD
        if fill: cell.fill = fill
        # 占同比条件着色(覆盖普通fill)
        if col in ztb_idx and isinstance(val,(int,float)):
            if val >= ZTB_THRESH:
                cell.fill = GREEN; cell.font = GREEN_FONT if not bold else Font(bold=True,color='006100')
            elif val <= -ZTB_THRESH:
                cell.fill = RED; cell.font = RED_FONT if not bold else Font(bold=True,color='9C0006')

def set_widths(ws, first_w, rest_w, ncol, first_cols=1):
    for i in range(1, first_cols+1):
        ws.column_dimensions[get_column_letter(i)].width = first_w
    for c in range(first_cols+1, ncol+1):
        ws.column_dimensions[get_column_letter(c)].width = rest_w
